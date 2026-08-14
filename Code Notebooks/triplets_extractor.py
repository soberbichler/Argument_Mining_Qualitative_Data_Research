"""
triplets_extractor.py
"""

import re

import pandas as pd
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── ✏️  CONFIGURE HERE ────────────────────────────────────────────────────────
FILEPATH  = "/content/training_2.xlsx"
ARG_COL   = "arguments"
LIMIT     = 20
# ─────────────────────────────────────────────────────────────────────────────

MODEL_ID = "oberbics/llama-3.1-base-kg-extraction-full"

ONTOLOGY_TEXT = """Valid triples (you MUST prioritize these):
- subject: aid → object: damage | predicate: insufficient for
- subject: aid → object: recovery | predicate: insufficient for
- subject: aid → object: refugees | predicate: fails to reach
- subject: aid → object: victims | predicate: saved
- subject: austria → object: aid | predicate: delivers
- subject: damage → object: aid | predicate: limits capacity of
- subject: damage → object: aid | predicate: necessitates
- subject: military → object: recovery | predicate: enables
- subject: society → object: aid | predicate: is obligated to
- subject: society → object: aid | predicate: is requested to
- subject: society → object: aid | predicate: obligates to
- subject: society → object: aid | predicate: lost interest for
- subject: aid → object: victims | predicate: reaches
- subject: austria → object: heroic | predicate: acts
- subject: austria → object: aid | predicate: effective response to
- subject: damage → object: aid | predicate: limits capacity of
- subject: damage → object: aid | predicate: prevents
- subject: damage → object: banks | predicate: limited impact
- subject: damage → object: economy | predicate: benefits financially
- subject: damage → object: economy | predicate: causes long-term to
- subject: damage → object: germany | predicate: harms economy
- subject: damage → object: italy | predicate: harms economy
- subject: damage → object: recovery | predicate: delays
- subject: population density → object: damage | predicate: increases
- subject: damage → object: epidemic | predicate: risks
- subject: damage → object: disorder | predicate: causes
- subject: disorganization → object: aid | predicate: undermines
- subject: friendship → object: aid | predicate: obligates to
- subject: italian government → object: aid | predicate: delays
- subject: italian government → object: damage | predicate: effective response to
- subject: italian government → object: order | predicate: maintains
- subject: italian government → object: recovery | predicate: enables
- subject: italian government → object: recovery | predicate: insufficient for
- subject: italian government → object: italy | predicate: harms economy
- subject: disorder → object: military | predicate: justifies
- subject: military → object: aid | predicate: delays
- subject: military → object: aid | predicate: not used enough
- subject: military → object: recovery | predicate: enables
- subject: military → object: recovery | predicate: enables
- subject: italy → object: misinformation | predicate: spreads
- subject: italian government → object: order | predicate: maintains
- subject: aid → object: epidemic | predicate: prevents
- subject: damage → object: recovery | predicate: prevents
- subject: economy → object: recovery | predicate: enables
- subject: england → object: recovery | predicate: strategically supports
- subject: italian government → object: aid | predicate: enables
- subject: italian government → object: recovery | predicate: can efficiently handle
- subject: italian government → object: recovery | predicate: lacks plan
- subject: recovery → object: economy | predicate: hinders
- subject: aid → object: refugees | predicate: is requested for
- subject: refugees → object: aid | predicate: necessitates
- subject: refugees → object: recovery | predicate: enables
- subject: refugees → object: stereotype | predicate: challenges
- subject: royalty → object: italian government | predicate: demands action from
- subject: italian government → object: scientific | predicate: must implement
- subject: italian government → object: scientific | predicate: should implement
- subject: scientific → object: damage | predicate: explains
- subject: scientific → object: damage | predicate: prevents
- subject: scientific → object: future | predicate: predicts
- subject: society → object: damage | predicate: explains
- subject: society → object: aid | predicate: is requested to
- subject: damage → object: sympathy | predicate: evokes
- subject: sympathy → object: aid | predicate: motivates
- subject: trauma → object: verification | predicate: undermines
- subject: unity → object: recovery | predicate: enables
- subject: unity → object: aid | predicate: enables
- subject: unity → object: crisis | predicate: supports
- subject: messina → object: italian government | predicate: grateful for response
- subject: germany → object: aid | predicate: delivers
- subject: god → object: earthquake | predicate: not related
- subject: damage → object: trauma | predicate: causes
- subject: trauma → object: aid | predicate: undermines
- subject: military → object: order | predicate: maintains
- subject: military → object: victims | predicate: harms
- subject: damage → object: italian government | predicate: increases costs for
- subject: italian government → object: media | predicate: suppresses
- subject: society → object: italian government | predicate: protests
- subject: royalty → object: aid | predicate: supports
- subject: journalists → object: misinformation | predicate: spread
- subject: military → object: aid | predicate: abuses
- subject: misinformation → object: society | predicate: harms
- subject: society → object: aid | predicate: hinders
- subject: scientific → object: human ignorance | predicate: reveals
- subject: society → object: recovery | predicate: responsible for
- subject: damage → object: economy | predicate: harms
- subject: damage → object: germany | predicate: harms
- subject: italian government → object: misinformation | predicate: spreads
- subject: aid → object: organized | predicate: is well
- subject: aid → object: all cities | predicate: did not reach
- subject: international aid → object: political advantage | predicate: serves
- subject: international aid → object: recovery | predicate: not needed anymore for
- subject: material assistance → object: monetary assistance | predicate: more effective than
- subject: recovery → object: possible | predicate: not
- subject: damage → object: migration | predicate: necessitates
- subject: messina → object: recover | predicate: will
- subject: aid from germany → object: recovery | predicate: enables
- subject: aid from germany → object: superior | predicate: is
- subject: crisis management of italy → object: not working | predicate: is
- subject: disorganization → object: Italians inefficiency for crisis | predicate: proofs
- subject: disorganization → object: italian reputation | predicate: harms
- subject: italy → object: misinformation | predicate: fights
- subject: scientific → object: past earthquakes | predicate: compares to
- subject: aid → object: aid | predicate: necessitates
- subject: damage → object: recovery | predicate: hinders
- subject: damage → object: trauma | predicate: creates
- subject: inaction → object: aid | predicate: delays
- subject: messina → object: aid | predicate: delays
- subject: italian government → object: narrative | predicate: controls
- subject: italian government → object: misinformation | predicate: spread
- subject: italian government → object: recovery | predicate: delays
- subject: italian government → object: recovery | predicate: must prioritize
- subject: italian government → object: recovery | predicate: lacks money
- subject: royalty → object: aid | predicate: enables
- subject: royalty → object: recovery | predicate: enables
- subject: military → object: aid | predicate: enables
- subject: journalists → object: disorganization | predicate: reveal
- subject: society → object: italian government | predicate: protests
- subject: society → object: aid | predicate: obligates to
- subject: aid → object: italy | predicate: misused by
- subject: damage → object: navigation | predicate: prevents"""

VALID_TRIPLES = [
    {"subject": "aid",                       "predicate": "insufficient for",        "object": "damage"},
    {"subject": "aid",                       "predicate": "insufficient for",        "object": "recovery"},
    {"subject": "aid",                       "predicate": "fails to reach",          "object": "refugees"},
    {"subject": "aid",                       "predicate": "saved",                   "object": "victims"},
    {"subject": "austria",                   "predicate": "delivers",                "object": "aid"},
    {"subject": "damage",                    "predicate": "limits capacity of",      "object": "aid"},
    {"subject": "damage",                    "predicate": "necessitates",            "object": "aid"},
    {"subject": "military",                  "predicate": "enables",                 "object": "recovery"},
    {"subject": "society",                   "predicate": "is obligated to",         "object": "aid"},
    {"subject": "society",                   "predicate": "is requested to",         "object": "aid"},
    {"subject": "society",                   "predicate": "obligates to",            "object": "aid"},
    {"subject": "society",                   "predicate": "lost interest for",       "object": "aid"},
    {"subject": "aid",                       "predicate": "reaches",                 "object": "victims"},
    {"subject": "austria",                   "predicate": "acts",                    "object": "heroic"},
    {"subject": "austria",                   "predicate": "effective response to",   "object": "aid"},
    {"subject": "damage",                    "predicate": "limits capacity of",      "object": "aid"},
    {"subject": "damage",                    "predicate": "prevents",                "object": "aid"},
    {"subject": "damage",                    "predicate": "limited impact",          "object": "banks"},
    {"subject": "damage",                    "predicate": "benefits financially",    "object": "economy"},
    {"subject": "damage",                    "predicate": "causes long-term to",     "object": "economy"},
    {"subject": "damage",                    "predicate": "harms economy",           "object": "germany"},
    {"subject": "damage",                    "predicate": "harms economy",           "object": "italy"},
    {"subject": "damage",                    "predicate": "delays",                  "object": "recovery"},
    {"subject": "population density",        "predicate": "increases",               "object": "damage"},
    {"subject": "damage",                    "predicate": "risks",                   "object": "epidemic"},
    {"subject": "damage",                    "predicate": "causes",                  "object": "disorder"},
    {"subject": "disorganization",           "predicate": "undermines",              "object": "aid"},
    {"subject": "friendship",                "predicate": "obligates to",            "object": "aid"},
    {"subject": "italian government",        "predicate": "delays",                  "object": "aid"},
    {"subject": "italian government",        "predicate": "effective response to",   "object": "damage"},
    {"subject": "italian government",        "predicate": "maintains",               "object": "order"},
    {"subject": "italian government",        "predicate": "enables",                 "object": "recovery"},
    {"subject": "italian government",        "predicate": "insufficient for",        "object": "recovery"},
    {"subject": "italian government",        "predicate": "harms economy",           "object": "italy"},
    {"subject": "disorder",                  "predicate": "justifies",               "object": "military"},
    {"subject": "military",                  "predicate": "delays",                  "object": "aid"},
    {"subject": "military",                  "predicate": "not used enough",         "object": "aid"},
    {"subject": "military",                  "predicate": "enables",                 "object": "recovery"},
    {"subject": "italy",                     "predicate": "spreads",                 "object": "misinformation"},
    {"subject": "aid",                       "predicate": "prevents",                "object": "epidemic"},
    {"subject": "damage",                    "predicate": "prevents",                "object": "recovery"},
    {"subject": "economy",                   "predicate": "enables",                 "object": "recovery"},
    {"subject": "england",                   "predicate": "strategically supports",  "object": "recovery"},
    {"subject": "italian government",        "predicate": "enables",                 "object": "aid"},
    {"subject": "italian government",        "predicate": "can efficiently handle",  "object": "recovery"},
    {"subject": "italian government",        "predicate": "lacks plan",              "object": "recovery"},
    {"subject": "recovery",                  "predicate": "hinders",                 "object": "economy"},
    {"subject": "aid",                       "predicate": "is requested for",        "object": "refugees"},
    {"subject": "refugees",                  "predicate": "necessitates",            "object": "aid"},
    {"subject": "refugees",                  "predicate": "enables",                 "object": "recovery"},
    {"subject": "refugees",                  "predicate": "challenges",              "object": "stereotype"},
    {"subject": "royalty",                   "predicate": "demands action from",     "object": "italian government"},
    {"subject": "italian government",        "predicate": "must implement",          "object": "scientific"},
    {"subject": "italian government",        "predicate": "should implement",        "object": "scientific"},
    {"subject": "scientific",                "predicate": "explains",                "object": "damage"},
    {"subject": "scientific",                "predicate": "prevents",                "object": "damage"},
    {"subject": "scientific",                "predicate": "predicts",                "object": "future"},
    {"subject": "society",                   "predicate": "explains",                "object": "damage"},
    {"subject": "society",                   "predicate": "is requested to",         "object": "aid"},
    {"subject": "damage",                    "predicate": "evokes",                  "object": "sympathy"},
    {"subject": "sympathy",                  "predicate": "motivates",               "object": "aid"},
    {"subject": "trauma",                    "predicate": "undermines",              "object": "verification"},
    {"subject": "unity",                     "predicate": "enables",                 "object": "recovery"},
    {"subject": "unity",                     "predicate": "enables",                 "object": "aid"},
    {"subject": "unity",                     "predicate": "supports",                "object": "crisis"},
    {"subject": "messina",                   "predicate": "grateful for response",   "object": "italian government"},
    {"subject": "germany",                   "predicate": "delivers",                "object": "aid"},
    {"subject": "god",                       "predicate": "not related",             "object": "earthquake"},
    {"subject": "damage",                    "predicate": "causes",                  "object": "trauma"},
    {"subject": "trauma",                    "predicate": "undermines",              "object": "aid"},
    {"subject": "military",                  "predicate": "maintains",               "object": "order"},
    {"subject": "military",                  "predicate": "harms",                   "object": "victims"},
    {"subject": "damage",                    "predicate": "increases costs for",     "object": "italian government"},
    {"subject": "italian government",        "predicate": "suppresses",              "object": "media"},
    {"subject": "society",                   "predicate": "protests",                "object": "italian government"},
    {"subject": "royalty",                   "predicate": "supports",                "object": "aid"},
    {"subject": "journalists",               "predicate": "spread",                  "object": "misinformation"},
    {"subject": "military",                  "predicate": "abuses",                  "object": "aid"},
    {"subject": "misinformation",            "predicate": "harms",                   "object": "society"},
    {"subject": "society",                   "predicate": "hinders",                 "object": "aid"},
    {"subject": "scientific",                "predicate": "reveals",                 "object": "human ignorance"},
    {"subject": "society",                   "predicate": "responsible for",         "object": "recovery"},
    {"subject": "damage",                    "predicate": "harms",                   "object": "economy"},
    {"subject": "damage",                    "predicate": "harms",                   "object": "germany"},
    {"subject": "italian government",        "predicate": "spreads",                 "object": "misinformation"},
    {"subject": "aid",                       "predicate": "is well",                 "object": "organized"},
    {"subject": "aid",                       "predicate": "did not reach",           "object": "all cities"},
    {"subject": "international aid",         "predicate": "serves",                  "object": "political advantage"},
    {"subject": "international aid",         "predicate": "not needed anymore for",  "object": "recovery"},
    {"subject": "material assistance",       "predicate": "more effective than",     "object": "monetary assistance"},
    {"subject": "recovery",                  "predicate": "not",                     "object": "possible"},
    {"subject": "damage",                    "predicate": "necessitates",            "object": "migration"},
    {"subject": "messina",                   "predicate": "will",                    "object": "recover"},
    {"subject": "aid from germany",          "predicate": "enables",                 "object": "recovery"},
    {"subject": "aid from germany",          "predicate": "is",                      "object": "superior"},
    {"subject": "crisis management of italy","predicate": "is",                      "object": "not working"},
    {"subject": "disorganization",           "predicate": "proofs",                  "object": "Italians inefficiency for crisis"},
    {"subject": "disorganization",           "predicate": "harms",                   "object": "italian reputation"},
    {"subject": "italy",                     "predicate": "fights",                  "object": "misinformation"},
    {"subject": "scientific",                "predicate": "compares to",             "object": "past earthquakes"},
    {"subject": "aid",                       "predicate": "necessitates",            "object": "aid"},
    {"subject": "damage",                    "predicate": "hinders",                 "object": "recovery"},
    {"subject": "damage",                    "predicate": "creates",                 "object": "trauma"},
    {"subject": "inaction",                  "predicate": "delays",                  "object": "aid"},
    {"subject": "messina",                   "predicate": "delays",                  "object": "aid"},
    {"subject": "italian government",        "predicate": "controls",                "object": "narrative"},
    {"subject": "italian government",        "predicate": "spread",                  "object": "misinformation"},
    {"subject": "italian government",        "predicate": "delays",                  "object": "recovery"},
    {"subject": "italian government",        "predicate": "must prioritize",         "object": "recovery"},
    {"subject": "italian government",        "predicate": "lacks money",             "object": "recovery"},
    {"subject": "royalty",                   "predicate": "enables",                 "object": "aid"},
    {"subject": "royalty",                   "predicate": "enables",                 "object": "recovery"},
    {"subject": "military",                  "predicate": "enables",                 "object": "aid"},
    {"subject": "journalists",               "predicate": "reveal",                  "object": "disorganization"},
    {"subject": "society",                   "predicate": "protests",                "object": "italian government"},
    {"subject": "society",                   "predicate": "obligates to",            "object": "aid"},
    {"subject": "aid",                       "predicate": "misused by",              "object": "italy"},
    {"subject": "damage",                    "predicate": "prevents",                "object": "navigation"},
]

# ── UPDATED: now allows the model to propose a triple outside the ontology ──
INSTRUCTION = (
    "Extract the knowledge graph triple from the argument below.\n\n"
    "ONTOLOGY — prioritize these triples:\n\n"
    f"{ONTOLOGY_TEXT}\n\n"
    "RULES:\n"
    "1. If the argument matches a triple from the ontology above, use it exactly as listed.\n"
    "2. If the argument contains a valid knowledge graph relation NOT in the ontology, suggest a NEW "
    "triple using the same format (subject, predicate, object). Set 'new_triple' to true.\n"
    "3. If no argument or relation exists, return NA for subject, predicate, object.\n\n"
    "Return your answer as JSON with keys: subject, predicate, object, confidence, new_triple.\n"
    "- confidence is a float between 0.0 and 1.0 indicating how well the triple fits.\n"
    "- new_triple is true if this triple is NOT in the ontology above, false otherwise."
)

OUTPUT_COLUMNS = ["kg_subject", "kg_predicate", "kg_object", "kg_confidence", "kg_valid", "kg_new_triple"]


def extract_fields(text):
    """Pull subject/predicate/object/confidence/new_triple out of the raw text.
    If the model returned multiple triples separated by '|', only the
    first is used."""
    first_segment = text.split("|")[0]

    def grab(key, segment):
        m = re.search(rf'"{key}"\s*:\s*"([^"]*)"', segment, re.IGNORECASE)
        if m:
            return m.group(1).strip()
        m2 = re.search(rf'"{key}"\s*:\s*([0-9.]+)', segment, re.IGNORECASE)
        if m2:
            return m2.group(1).strip()
        m3 = re.search(rf'"{key}"\s*:\s*(true|false)', segment, re.IGNORECASE)
        return m3.group(1).strip() if m3 else None

    subject   = grab("subject", first_segment)
    predicate = grab("predicate", first_segment)
    obj       = grab("object", first_segment)
    confidence_raw = grab("confidence", first_segment)
    new_triple_raw = grab("new_triple", first_segment)

    if subject is None and predicate is None and obj is None:
        return None

    try:
        confidence = float(confidence_raw) if confidence_raw else 0.0
    except ValueError:
        confidence = 0.0

    new_triple = str(new_triple_raw).strip().lower() == "true" if new_triple_raw is not None else False

    return {
        "subject": subject,
        "predicate": predicate,
        "object": obj,
        "confidence": confidence,
        "new_triple": new_triple,
    }


def validate(parsed):
    if parsed is None:
        return False, 0.0
    if str(parsed.get("subject", "")).upper() == "NA":
        return True, parsed.get("confidence", 0.0)
    if parsed.get("new_triple", False):
        return True, parsed.get("confidence", 0.0)
    subject   = parsed.get("subject")
    predicate = parsed.get("predicate")
    obj       = parsed.get("object")
    for t in VALID_TRIPLES:
        if t["subject"] == subject and t["predicate"] == predicate and t["object"] == obj:
            return True, parsed.get("confidence", 1.0)
    return False, parsed.get("confidence", 0.0)


def load_model():
    print("Loading model …")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_ID)
    tokenizer.pad_token = tokenizer.eos_token
    tokenizer.padding_side = "right"

    model = AutoModelForCausalLM.from_pretrained(
        MODEL_ID,
        torch_dtype=torch.float16,
        device_map="auto"
    )
    model.eval()

    n_instr_tokens = len(tokenizer(INSTRUCTION)["input_ids"])
    print(f" Model loaded (instruction block: {n_instr_tokens} tokens)")
    if n_instr_tokens > 1700:
        print("  WARNING: instruction is close to/over the 2048 token max_length "
              "used during training. Longer arguments may have been truncated "
              "during training, which can degrade output quality.")
    return model, tokenizer


def extract_kg(text, model, tokenizer):
    prompt = f"### Instruction:\n{INSTRUCTION}\n\n### Input:\n{text}\n\n### Response:\n"
    inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=180,
            do_sample=False,
            pad_token_id=tokenizer.eos_token_id,
        )
    generated = outputs[0][inputs["input_ids"].shape[1]:]
    raw = tokenizer.decode(generated, skip_special_tokens=True).strip()
    parsed = extract_fields(raw)
    valid, confidence = validate(parsed)
    return parsed, valid, confidence


def auto_detect_arg_col(df):
    candidates = [c for c in df.columns if "arg" in c.lower()]
    if candidates:
        return candidates[0]
    text_cols = df.select_dtypes(include="object").columns
    if len(text_cols) == 0:
        return df.columns[0]
    return max(text_cols, key=lambda c: df[c].astype(str).str.len().mean())


def main():
    print(f"Reading Sheet2 from: {FILEPATH}")
    df = pd.read_excel(FILEPATH, sheet_name="Sheet1")
    print(f"  {len(df)} rows, columns: {list(df.columns)}")

    arg_col = ARG_COL or auto_detect_arg_col(df)
    print(f"  Using argument column: '{arg_col}'")

    if LIMIT:
        df = df[42:82]
        print(f"  Limited to {LIMIT} rows for testing")

    model, tokenizer = load_model()

    results = []
    new_triples_found = []

    for i, row in df.iterrows():
        text = str(row[arg_col])
        print(f"  Row {i+1}/{len(df)} ...", end=" ", flush=True)

        if not text.strip() or text.strip().lower() in ("nan", "none", ""):
            results.append({"kg_subject": "NA", "kg_predicate": "NA",
                             "kg_object": "NA",
                             "kg_confidence": 0.0, "kg_valid": False,
                             "kg_new_triple": False})
            print("(empty - skipped)")
            continue

        parsed, valid, confidence = extract_kg(text, model, tokenizer)

        if parsed:
            is_new = parsed.get("new_triple", False)
            result = {
                "kg_subject":    parsed.get("subject",    "NA"),
                "kg_predicate":  parsed.get("predicate",  "NA"),
                "kg_object":     parsed.get("object",     "NA"),
                "kg_confidence": confidence,
                "kg_valid":      valid,
                "kg_new_triple": is_new,
            }
            results.append(result)
            if is_new:
                new_triples_found.append({
                    "row": i+1,
                    "subject":   parsed.get("subject"),
                    "predicate": parsed.get("predicate"),
                    "object":    parsed.get("object"),
                    "input":     text[:100]
                })
                print(f"NEW  {result['kg_subject']} -> {result['kg_object']}  (suggested new triple)")
            else:
                print(f"OK   {result['kg_subject']} -> {result['kg_object']}  (valid={valid})")
        else:
            results.append({"kg_subject": "PARSE_ERROR", "kg_predicate": "PARSE_ERROR",
                             "kg_object": "PARSE_ERROR",
                             "kg_confidence": 0.0, "kg_valid": False, "kg_new_triple": False})
            print("PARSE_ERROR")

    if new_triples_found:
        print(f"\n=== {len(new_triples_found)} NEW TRIPLES SUGGESTED ===")
        for t in new_triples_found:
            print(f"  Row {t['row']}: {t['subject']} | {t['predicate']} | {t['object']}")

    print("\nWriting results back to Sheet2 ...")
    wb = load_workbook(FILEPATH)
    ws = wb["Sheet1"]

    start_col = ws.max_column + 1
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    for j, col_name in enumerate(OUTPUT_COLUMNS):
        cell = ws.cell(row=1, column=start_col + j, value=col_name)
        cell.font = Font(bold=True, name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, result in enumerate(results):
        excel_row = row_idx + 2
        for j, col_name in enumerate(OUTPUT_COLUMNS):
            val = result.get(col_name)
            cell = ws.cell(row=excel_row, column=start_col + j, value=val)
            cell.font = Font(name="Arial")
            if col_name == "kg_valid":
                color = "C6EFCE" if val else "FFC7CE"
                cell.fill = PatternFill("solid", fgColor=color)
            if col_name == "kg_new_triple" and val:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")

    for j, col_name in enumerate(OUTPUT_COLUMNS):
        col_letter = get_column_letter(start_col + j)
        ws.column_dimensions[col_letter].width = max(
            len(col_name) + 2,
            max((len(str(r.get(col_name, ""))) for r in results), default=10) + 2
        )

    wb.save(FILEPATH)
    print(f"Done! Results written to Sheet2 starting at column {get_column_letter(start_col)}")
    print(f"   Valid triples: {sum(r['kg_valid'] for r in results)}/{len(results)}")
    print(f"   New triples suggested: {sum(r['kg_new_triple'] for r in results)}")


if __name__ == "__main__":
    main()
